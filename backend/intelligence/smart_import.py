from __future__ import annotations

import csv
import io
import json
import logging
import os
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    from google import genai

    GEMINI_SDK_AVAILABLE = True
except ImportError:
    genai = None
    GEMINI_SDK_AVAILABLE = False

try:
    from openpyxl import load_workbook

    OPENPYXL_AVAILABLE = True
except ImportError:
    load_workbook = None
    OPENPYXL_AVAILABLE = False

from backend.intelligence.trade_entry_parser import (
    GEMINI_MODEL,
    PROVIDER,
    _extract_response_text,
    _loads_json_object,
)

logger = logging.getLogger(__name__)

MAX_FILE_BYTES = 5 * 1024 * 1024
BULLCAST_FIELDS = [
    "date",
    "symbol",
    "side",
    "entry",
    "exit",
    "quantity",
    "setup_tag",
    "mistake_tag",
    "confidence_score",
    "notes",
    "planned_risk",
    "planned_reward",
]
OPTIONAL_TEXT_FIELDS = {"setup_tag", "mistake_tag", "notes"}
NUMBER_FIELDS = {"entry", "exit", "quantity", "confidence_score", "planned_risk", "planned_reward"}
MISSING_KEY_WARNING = "Gemini smart import unavailable because GEMINI_API_KEY is not configured."
MISSING_SDK_WARNING = "Gemini smart import unavailable because google-genai is not installed."
REQUEST_FAILED_WARNING = "Gemini smart import unavailable because the Gemini request failed."
INVALID_JSON_WARNING = "Gemini smart import returned invalid JSON."
DETERMINISTIC_FALLBACK_WARNING = "Gemini unavailable — column mapping produced by deterministic heuristic fallback."
DETERMINISTIC_FALLBACK_ORIGIN = "smart_import_deterministic_fallback"
GEMINI_FILE_IMPORT_ORIGIN = "gemini_file_import"
DETERMINISTIC_COLUMN_MAP: dict[str, str] = {
    "symbol": "symbol",
    "ticker": "symbol",
    "stock": "symbol",
    "instrument": "symbol",
    "scrip": "symbol",
    "date": "date",
    "trade date": "date",
    "time": "date",
    "datetime": "date",
    "side": "side",
    "direction": "side",
    "type": "side",
    "action": "side",
    "buy or sell": "side",
    "buy sell": "side",
    "entry": "entry",
    "buy price": "entry",
    "open price": "entry",
    "avg buy": "entry",
    "average buy": "entry",
    "filled": "entry",
    "fill price": "entry",
    "entry price": "entry",
    "exit": "exit",
    "sell price": "exit",
    "close price": "exit",
    "avg sell": "exit",
    "average sell": "exit",
    "exit price": "exit",
    "qty": "quantity",
    "quantity": "quantity",
    "shares": "quantity",
    "lots": "quantity",
    "units": "quantity",
    "size": "quantity",
    "setup": "setup_tag",
    "strategy": "setup_tag",
    "pattern": "setup_tag",
    "signal": "setup_tag",
    "mistake": "mistake_tag",
    "error": "mistake_tag",
    "violation": "mistake_tag",
    "confidence": "confidence_score",
    "score": "confidence_score",
    "rating": "confidence_score",
    "notes": "notes",
    "comment": "notes",
    "remarks": "notes",
    "journal": "notes",
    "risk": "planned_risk",
    "planned risk": "planned_risk",
    "stop loss distance": "planned_risk",
    "reward": "planned_reward",
    "planned reward": "planned_reward",
    "target distance": "planned_reward",
}


class GeminiSmartImportError(Exception):
    def __init__(self, warning: str):
        super().__init__(warning)
        self.warning = warning


def parse_uploaded_file(file_bytes: bytes, filename: str) -> tuple[list[str], list[dict[str, Any]]]:
    if len(file_bytes or b"") > MAX_FILE_BYTES:
        raise ValueError("Smart Import files must be 5MB or smaller.")

    suffix = Path(str(filename or "")).suffix.lower()
    if suffix == ".csv":
        return _parse_csv(file_bytes)
    if suffix == ".xlsx":
        return _parse_xlsx(file_bytes)
    raise ValueError("Smart Import accepts only .csv and .xlsx files.")


def get_column_mapping(headers: list[str], sample_rows: list[dict[str, Any]]) -> tuple[dict[str, str | None], bool]:
    clean_headers = [str(header or "").strip() for header in headers if str(header or "").strip()]
    if not clean_headers:
        return {}, False

    api_key = os.getenv("GEMINI_API_KEY", "").strip()
    if not api_key:
        logger.warning(MISSING_KEY_WARNING)
        return get_deterministic_column_mapping(clean_headers), True
    if not GEMINI_SDK_AVAILABLE or genai is None:
        logger.warning(MISSING_SDK_WARNING)
        return get_deterministic_column_mapping(clean_headers), True

    try:
        client = genai.Client(api_key=api_key)
        response = client.models.generate_content(
            model=GEMINI_MODEL,
            contents=_build_mapping_prompt(clean_headers, sample_rows[:5]),
        )
        response_text = _extract_response_text(response)
        if not response_text:
            raise GeminiSmartImportError(INVALID_JSON_WARNING)
        try:
            payload = _loads_json_object(response_text)
        except ValueError as exc:
            raise GeminiSmartImportError(INVALID_JSON_WARNING) from exc
        return _sanitize_mapping(payload, clean_headers), False
    except GeminiSmartImportError:
        logger.warning("Gemini smart import mapping failed; using deterministic fallback")
        return get_deterministic_column_mapping(clean_headers), True
    except Exception:
        logger.warning("Gemini smart import mapping failed")
        return get_deterministic_column_mapping(clean_headers), True


def get_deterministic_column_mapping(headers: list[str]) -> dict[str, str | None]:
    mapping: dict[str, str | None] = {}
    for header in headers:
        source = str(header or "").strip()
        mapping[source] = DETERMINISTIC_COLUMN_MAP.get(_normalize_column_name(source))
    return mapping


def apply_mapping(
    rows: list[dict[str, Any]],
    mapping: dict[str, str | None],
    *,
    data_origin: str = GEMINI_FILE_IMPORT_ORIGIN,
) -> list[dict[str, Any]]:
    field_to_source: dict[str, str] = {}
    for source, target in (mapping or {}).items():
        clean_target = _normalize_target_field(target)
        if clean_target and clean_target not in field_to_source:
            field_to_source[clean_target] = str(source)

    trades: list[dict[str, Any]] = []
    for row in rows:
        trade: dict[str, Any] = {
            "date": _safe_iso_date(_mapped_value(row, field_to_source, "date")),
            "symbol": _safe_text(_mapped_value(row, field_to_source, "symbol")).upper() or None,
            "side": _normalize_side(_mapped_value(row, field_to_source, "side")),
            "entry": _safe_number(_mapped_value(row, field_to_source, "entry")),
            "exit": _safe_number(_mapped_value(row, field_to_source, "exit")),
            "quantity": _safe_number(_mapped_value(row, field_to_source, "quantity")),
            "setup": _safe_text(_mapped_value(row, field_to_source, "setup_tag")),
            "setup_tag": _normalize_tag(_mapped_value(row, field_to_source, "setup_tag")),
            "confidence": "",
            "confidence_score": _safe_confidence(_mapped_value(row, field_to_source, "confidence_score")),
            "mistake": _safe_text(_mapped_value(row, field_to_source, "mistake_tag")),
            "mistake_tag": _normalize_mistake_tag(_mapped_value(row, field_to_source, "mistake_tag")),
            "rule_followed": None,
            "planned_risk": _safe_number(_mapped_value(row, field_to_source, "planned_risk")),
            "planned_reward": _safe_number(_mapped_value(row, field_to_source, "planned_reward")),
            "entry_reason": "",
            "exit_reason": "",
            "notes": _safe_text(_mapped_value(row, field_to_source, "notes")),
            "data_origin": data_origin,
            "is_synthetic": False,
        }

        missing_fields = _missing_fields(trade, field_to_source)
        trade["needs_review"] = bool(missing_fields)
        trade["missing_fields"] = missing_fields
        trades.append(trade)

    return trades


def mapping_warnings(mapping: dict[str, str | None]) -> list[str]:
    mapped_fields = {target for target in (mapping or {}).values() if target in BULLCAST_FIELDS}
    missing = [field for field in BULLCAST_FIELDS if field not in mapped_fields]
    if not missing:
        return []
    return [f"Unmapped Bullcast fields: {', '.join(missing)}."]


def _parse_csv(file_bytes: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    text = _decode_csv(file_bytes)
    reader = csv.DictReader(io.StringIO(text))
    headers = [str(header or "").strip() for header in (reader.fieldnames or [])]
    rows: list[dict[str, Any]] = []
    for row in reader:
        rows.append({header: _cell_value(row.get(header)) for header in headers})
    return headers, rows


def _parse_xlsx(file_bytes: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    if not OPENPYXL_AVAILABLE or load_workbook is None:
        raise ValueError("XLSX Smart Import requires openpyxl.")
    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    try:
        header_row = next(iterator)
    except StopIteration:
        return [], []
    headers = [str(value or "").strip() for value in header_row]
    rows: list[dict[str, Any]] = []
    for values in iterator:
        row: dict[str, Any] = {}
        for index, header in enumerate(headers):
            row[header] = _cell_value(values[index] if index < len(values) else None)
        if any(value not in ("", None) for value in row.values()):
            rows.append(row)
    return headers, rows


def _decode_csv(file_bytes: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return file_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    return file_bytes.decode("utf-8", errors="replace")


def _normalize_column_name(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").strip().lower()).strip()


def _build_mapping_prompt(headers: list[str], sample_rows: list[dict[str, Any]]) -> str:
    schema = {
        "column_mapping": {header: "date | symbol | side | entry | exit | quantity | setup_tag | mistake_tag | confidence_score | notes | planned_risk | planned_reward | null" for header in headers}
    }
    return (
        "You are a column mapper for Bullcast trade journal imports.\n"
        "Map each source column name to the closest Bullcast schema field, or null if there is no confident match.\n"
        "This is column mapping only. Do not rewrite, normalize, infer, or invent any cell values.\n"
        "Do not provide trading advice. Return valid JSON only.\n"
        f"Allowed Bullcast fields: {BULLCAST_FIELDS}.\n"
        "Return this exact top-level shape:\n"
        f"{json.dumps(schema, ensure_ascii=True)}\n\n"
        f"Source headers: {json.dumps(headers, ensure_ascii=True)}\n"
        f"First sample rows: {json.dumps(sample_rows[:5], ensure_ascii=True, default=str)}"
    )


def _sanitize_mapping(payload: dict[str, Any], headers: list[str]) -> dict[str, str | None]:
    raw_mapping = payload.get("column_mapping") if isinstance(payload.get("column_mapping"), dict) else payload
    mapping: dict[str, str | None] = {}
    for header in headers:
        mapping[header] = _normalize_target_field(raw_mapping.get(header))
    return mapping


def _normalize_target_field(value: Any) -> str | None:
    if value in ("", None):
        return None
    text = str(value).strip().lower().replace(" ", "_").replace("-", "_")
    aliases = {
        "ticker": "symbol",
        "trade_date": "date",
        "direction": "side",
        "type": "side",
        "buy_price": "entry",
        "entry_price": "entry",
        "sell_price": "exit",
        "exit_price": "exit",
        "shares": "quantity",
        "qty": "quantity",
        "strategy": "setup_tag",
        "setup": "setup_tag",
    }
    text = aliases.get(text, text)
    return text if text in BULLCAST_FIELDS else None


def _mapped_value(row: dict[str, Any], field_to_source: dict[str, str], field: str) -> Any:
    source = field_to_source.get(field)
    return row.get(source) if source else None


def _missing_fields(trade: dict[str, Any], field_to_source: dict[str, str]) -> list[str]:
    missing: list[str] = []
    for field in BULLCAST_FIELDS:
        if field not in field_to_source:
            missing.append(field)
            continue
        value = trade.get(field)
        if field in OPTIONAL_TEXT_FIELDS:
            continue
        if value is None or value == "":
            missing.append(field)
    return missing


def _cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()[:10]
    if isinstance(value, str):
        return value.strip()
    return value


def _safe_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _safe_number(value: Any) -> float | None:
    if value in ("", None):
        return None
    if isinstance(value, str):
        value = value.replace(",", "").replace("$", "").replace("%", "").strip()
        if not value:
            return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if number != number or number in {float("inf"), float("-inf")}:
        return None
    return number


def _safe_confidence(value: Any) -> int | None:
    number = _safe_number(value)
    if number is None:
        return None
    integer = int(number)
    return integer if 1 <= integer <= 5 else None


def _safe_iso_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    try:
        return date.fromisoformat(text[:10]).isoformat()
    except ValueError:
        return None


def _normalize_side(value: Any) -> str | None:
    text = str(value or "").strip().upper()
    if text in {"LONG", "BUY", "BOUGHT", "B"}:
        return "LONG"
    if text in {"SHORT", "SELL", "SOLD", "S"}:
        return "SHORT"
    return None


def _normalize_tag(value: Any) -> str:
    text = str(value or "").strip().lower().replace("-", "_").replace(" ", "_")
    if "breakout" in text:
        return "breakout"
    if "momentum" in text:
        return "momentum"
    if "pullback" in text or "retest" in text:
        return "pullback"
    if "reversal" in text:
        return "reversal"
    if "gap" in text:
        return "news_event"
    return text[:64]


def _normalize_mistake_tag(value: Any) -> str:
    text = str(value or "").strip().lower().replace("-", "_").replace(" ", "_")
    return text[:64] if text else "none"
